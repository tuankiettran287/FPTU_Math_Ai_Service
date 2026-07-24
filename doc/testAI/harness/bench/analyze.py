"""Tổng hợp số liệu -> Excel + biểu đồ + summary.json.

  python -m bench.analyze
"""
from __future__ import annotations

import json
import re
import statistics as st
from collections import Counter
from pathlib import Path

import numpy as np

from . import config, tasks

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "results"
REP = ROOT / "report"
REP.mkdir(exist_ok=True)
(REP / "charts").mkdir(exist_ok=True)


def pct(x, n):
    return round(100.0 * x / n, 2) if n else 0.0


def pk(vals, q):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return round(float(np.percentile(vals, q)), 2)


def load(model, mode):
    p = OUT / f"graded_{model}_{mode}.jsonl"
    if not p.exists():
        return []
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]


# Chữ Hán/Nhật/Hàn — dùng để bắt hiện tượng model rò tiếng Trung vào câu trả lời tiếng Việt.
_CJK = re.compile(r"[一-鿿぀-ヿ가-힯]")


def cjk_leak(model: str, mode: str) -> dict:
    """Tỉ lệ câu trả lời bị lẫn chữ Trung/Nhật/Hàn.

    Vì sao chỉ số này đáng lên báo cáo: DeepSeek-R1-Distill được chưng cất từ dữ liệu
    đa ngữ nặng tiếng Trung, nên hay chèn chữ Hán vào giữa câu tiếng Việt
    (bắt được thật: "có thể他们是 đã tính sai số"). Với sản phẩm phục vụ sinh viên
    Việt Nam thì đây là lỗi chất lượng thấy được ngay, không phải tiểu tiết học thuật.
    Tách riêng phần <think> và phần ĐÁP ÁN: rò trong lúc 'nháp' thì người dùng không
    thấy, nhưng rò vào ĐÁP ÁN thì sinh viên đọc thẳng -> nặng hơn hẳn.
    """
    p = OUT / f"texts_{model}_{mode}.jsonl"
    if not p.exists():
        return {}
    n = leak_all = leak_ans = 0
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        t = json.loads(line).get("text") or ""
        if not t:
            continue
        n += 1
        if _CJK.search(t):
            leak_all += 1
        i = t.rfind("</think>")
        ans = t[i + 8:] if i != -1 else t
        if _CJK.search(ans):
            leak_ans += 1
    return {"cjk_leak_rate": pct(leak_all, n),
            "cjk_leak_answer_rate": pct(leak_ans, n),
            "cjk_n": n}


# ── Context Relevance đo bằng embedding (khách quan, KHÔNG nhờ giám khảo) ─────
def context_relevance_embedding() -> dict[str, float]:
    """So ngữ cảnh truy hồi được với expected_context của bộ test.

    Đây là số liệu ĐỘC LẬP với giám khảo LLM: nếu vector của đoạn lấy về gần với
    mô tả 'kiến thức cần dùng' do người ra đề viết, thì retrieval đã lấy đúng.
    """
    from .client import embed
    items = tasks.load_test_set(ROOT / "testset")
    ctx = json.loads((OUT / "contexts.json").read_text(encoding="utf-8"))
    ids = [i["test_id"] for i in items if i["test_id"] in ctx]
    exp = [next(x for x in items if x["test_id"] == i)["expected_context"] for i in ids]
    out = {}
    B = 64
    for s in range(0, len(ids), B):
        vs = np.asarray(embed(exp[s:s + B], port=config.EMBED_PORT, model=config.EMBED_MODEL),
                        dtype=np.float32)
        vs /= np.linalg.norm(vs, axis=1, keepdims=True) + 1e-12
        chunk_ids = ids[s:s + B]
        hit_texts, spans = [], []
        for tid in chunk_ids:
            hs = ctx[tid]["hits"]
            spans.append((len(hit_texts), len(hit_texts) + len(hs)))
            hit_texts += [h["text"][:2000] for h in hs]
        if not hit_texts:
            continue
        hv = np.asarray(embed(hit_texts, port=config.EMBED_PORT, model=config.EMBED_MODEL),
                        dtype=np.float32)
        hv /= np.linalg.norm(hv, axis=1, keepdims=True) + 1e-12
        for k, tid in enumerate(chunk_ids):
            a, b = spans[k]
            if b > a:
                out[tid] = float((hv[a:b] @ vs[k]).max())
    (OUT / "context_relevance_emb.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    return out


def summarize(rows, cre: dict[str, float]) -> dict:
    n = len(rows)
    ok = [r for r in rows if r.get("ok")]
    d = {
        "n": n,
        "n_ok": len(ok),
        "error_rate": pct(n - len(ok), n),
        "timeout_rate": pct(sum(1 for r in rows if r.get("timeout")), n),
        "oom_rate": pct(sum(1 for r in rows if "memory" in (r.get("error") or "").lower()), n),
        "truncation_rate": pct(sum(1 for r in ok if r.get("truncated")), len(ok)),
        "format_compliance": pct(sum(1 for r in ok if r.get("format_ok")), len(ok)),
    }
    # Hiệu năng: CHỈ lấy từ pha conc=1 -> số sạch, không lẫn hàng đợi.
    perf = [r for r in ok if r.get("phase_kind") == "perf"]
    d["perf_n"] = len(perf)
    for name, key in (("ttft_ms", "ttft_ms"), ("latency_s", "total_latency_s"),
                      ("tokens_per_sec", "tokens_per_sec"), ("tpot_ms", "tpot_ms"),
                      ("prefill_s", "prefill_time_s"), ("decode_s", "decode_time_s")):
        vals = [r.get(key) for r in perf if r.get(key) is not None]
        d[f"{name}_p50"] = pk(vals, 50)
        d[f"{name}_p95"] = pk(vals, 95)
        d[f"{name}_p99"] = pk(vals, 99)
        d[f"{name}_mean"] = round(st.mean(vals), 2) if vals else None
    # Suy nghĩ (chỉ có ý nghĩa với model reasoning)
    tt = [r.get("think_time_s") for r in perf if r.get("think_time_s")]
    d["think_time_s_mean"] = round(st.mean(tt), 2) if tt else None
    ratios = [r["think_tokens"] / r["completion_tokens"]
              for r in ok if r.get("completion_tokens") and r.get("think_tokens") is not None]
    d["think_token_ratio"] = round(st.mean(ratios), 4) if ratios else None
    d["think_tokens_mean"] = round(st.mean([r.get("think_tokens", 0) for r in ok]), 1) if ok else None
    d["answer_tokens_mean"] = round(st.mean([r.get("answer_tokens", 0) for r in ok]), 1) if ok else None
    # Token & chi phí
    d["prompt_tokens_total"] = sum(r.get("prompt_tokens", 0) for r in ok)
    d["completion_tokens_total"] = sum(r.get("completion_tokens", 0) for r in ok)
    # Độ chính xác theo từng vấn đề
    for prob in ("V1", "V2", "V3", "V2H"):
        sub = [r for r in ok if r.get("problem") == prob and "correct" in r]
        d[f"acc_{prob}"] = pct(sum(1 for r in sub if r["correct"]), len(sub))
        d[f"n_{prob}"] = len(sub)
    v2 = [r for r in ok if r.get("problem") == "V2" and "correct" in r]
    for s in sorted({r["subject"] for r in v2}):
        g = [r for r in v2 if r["subject"] == s]
        d[f"acc_V2_{s}"] = pct(sum(1 for r in g if r["correct"]), len(g))
    for dif in ("easy", "medium", "hard"):
        g = [r for r in v2 if r.get("difficulty") == dif]
        d[f"acc_V2_{dif}"] = pct(sum(1 for r in g if r["correct"]), len(g))
    v2h = [r for r in ok if r.get("problem") == "V2H" and "correct" in r]
    for s_ in sorted({r["subject"] for r in v2h}):
        g = [r for r in v2h if r["subject"] == s_]
        d[f"acc_V2H_{s_}"] = pct(sum(1 for r in g if r["correct"]), len(g))
    for dif in ("medium", "hard"):
        g = [r for r in v2h if r.get("difficulty") == dif]
        if g:
            d[f"acc_V2H_{dif}"] = pct(sum(1 for r in g if r["correct"]), len(g))
    # Phân loại lỗi trên bộ KHÓ — mới là chỗ lỗi thật sự xuất hiện đủ nhiều để phân tích
    errs_h = Counter(r.get("error_category") for r in v2h if not r.get("correct"))
    tot_h = sum(errs_h.values())
    for k in ("retrieval", "calculation", "hallucination", "unknown", "none"):
        d[f"errH_{k}"] = pct(errs_h.get(k, 0), tot_h)
    d["n_errors_hard"] = tot_h

    # Phân loại lỗi (chỉ trên V2 sai)
    errs = Counter(r.get("error_category") for r in v2 if not r.get("correct"))
    tot_err = sum(errs.values())
    for k in ("retrieval", "calculation", "hallucination", "unknown", "none"):
        d[f"err_{k}"] = pct(errs.get(k, 0), tot_err)
    d["n_errors"] = tot_err
    # V3: ma trận nhầm lẫn — chỉ số GV quan tâm nhất
    v3 = [r for r in ok if r.get("problem") == "V3" and r.get("verdict")]
    if v3:
        fp = sum(1 for r in v3 if r["label"] == "incorrect" and r["verdict"] == "correct")
        fn = sum(1 for r in v3 if r["label"] == "correct" and r["verdict"] == "incorrect")
        d["v3_false_pass_rate"] = pct(fp, sum(1 for r in v3 if r["label"] == "incorrect"))
        d["v3_false_fail_rate"] = pct(fn, sum(1 for r in v3 if r["label"] == "correct"))
        for var in sorted({r.get("variant") for r in v3 if r.get("variant")}):
            g = [r for r in v3 if r.get("variant") == var]
            d[f"v3_acc_{var}"] = pct(sum(1 for r in g if r["correct"]), len(g))
    # Bộ ba RAG
    rag = [r for r in ok if r.get("mode") == "rag" and r.get("problem") == "V2"]
    if rag:
        cs = [r for r in rag if r.get("context_sufficient")]
        d["rag_context_relevance"] = pct(sum(1 for r in cs if r["context_sufficient"] == "yes"), len(cs))
        gr = [r for r in rag if r.get("grounded")]
        d["rag_groundedness"] = pct(sum(1 for r in gr if r["grounded"] in ("yes", "partial")), len(gr))
        d["rag_groundedness_strict"] = pct(sum(1 for r in gr if r["grounded"] == "yes"), len(gr))
        d["rag_answer_relevance"] = d["acc_V2"]
        d["rag_citation_rate"] = pct(sum(1 for r in rag if r.get("cited")), len(rag))
        sims = [cre[r["task_id"]] for r in rag if r["task_id"] in cre]
        d["rag_context_relevance_emb"] = round(st.mean(sims), 4) if sims else None
        cont = [r for r in rag if (r.get("max_exemplar_sim") or 0) >= config.CONTAMINATION_SIM]
        d["rag_contamination_rate"] = pct(len(cont), len(rag))
        d["rag_theory_hit_rate"] = pct(sum(1 for r in rag if r.get("n_theory_hits")), len(rag))
    return d


def costs(summ: dict) -> list[dict]:
    rows = []
    for (m, mode), d in summ.items():
        pin, pout = d["prompt_tokens_total"], d["completion_tokens_total"]
        r = {"model": m, "mode": mode, "prompt_tokens": pin, "completion_tokens": pout,
             "total_tokens": pin + pout, "local_usd": 0.0}
        for api, p in config.API_PRICING.items():
            r[f"{api}_usd"] = round(pin / 1e6 * p["in"] + pout / 1e6 * p["out"], 4)
        rows.append(r)
    return rows


def charts(summ: dict):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams.update({"figure.dpi": 140, "font.size": 9,
                         "axes.grid": True, "grid.alpha": .3})
    keys = [m.key for m in config.MODELS]
    labels = [m.short or m.label.split(" (")[0] for m in config.MODELS]
    C = {"pure": "#94a3b8", "rag": "#2563eb"}

    def grouped(metric, title, ylabel, fname, fmt="{:.1f}"):
        fig, ax = plt.subplots(figsize=(max(7, 1.9 * len(keys)), 3.7))
        x = np.arange(len(keys)); w = 0.36
        for i, mode in enumerate(("pure", "rag")):
            v = [(summ.get((k, mode), {}) or {}).get(metric) or 0 for k in keys]
            b = ax.bar(x + (i - .5) * w, v, w, label="Prompt thuần" if mode == "pure" else "Prompt + RAG",
                       color=C[mode])
            ax.bar_label(b, fmt=fmt, fontsize=7, padding=1)
        ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=8)
        ax.set_title(title); ax.set_ylabel(ylabel); ax.legend(fontsize=8)
        fig.tight_layout(); fig.savefig(REP / "charts" / fname); plt.close(fig)

    # Slide 3 — hiệu năng
    grouped("ttft_ms_p50", "Time To First Token (P50, concurrency=1)", "ms", "s3_ttft.png")
    grouped("tokens_per_sec_mean", "Tốc độ sinh token (concurrency=1)", "tokens/giây", "s3_speed.png")
    grouped("latency_s_p95", "Tổng thời gian trả lời (P95)", "giây", "s3_latency_p95.png")
    # VRAM: TÁCH trọng số (bất biến của model) khỏi KV cache (do mình cấu hình).
    # Gộp chung thành một cột "VRAM tiêu thụ" là gây hiểu sai: hai model 7B có trọng
    # số y hệt nhau nhưng nvidia-smi báo lệch 24GB chỉ vì util đặt khác.
    mem = {}
    p = OUT / "vllm_mem.json"
    if p.exists():
        mem = json.loads(p.read_text(encoding="utf-8"))
    fig, ax = plt.subplots(figsize=(max(7.6, 1.9 * len(keys)), 3.9))
    w = [mem.get(k, {}).get("weights_gib", 0) for k in keys]
    kv = [mem.get(k, {}).get("kv_cache_gib", 0) for k in keys]
    gr = [mem.get(k, {}).get("cuda_graph_gib", 0) for k in keys]
    b1 = ax.bar(labels, w, label="Trọng số model (bất biến)", color="#0e7490")
    b2 = ax.bar(labels, kv, bottom=w, label="KV cache (tuỳ cấu hình)", color="#67e8f9")
    ax.bar(labels, gr, bottom=[a + c for a, c in zip(w, kv)],
           label="CUDA graph", color="#cffafe")
    ax.bar_label(b1, fmt="%.1f", label_type="center", fontsize=7)
    ax.bar_label(b2, fmt="%.1f", label_type="center", fontsize=7)
    ax.axhline(95.6, ls="--", c="r", lw=1)
    ax.text(-0.4, 96.5, "Giới hạn RTX PRO 6000 — 96GB", fontsize=7, color="r")
    ax.set_title("VRAM: trọng số model vs KV cache")
    ax.set_ylabel("GiB"); ax.set_ylim(0, 105); ax.legend(fontsize=7, loc="upper left")
    fig.tight_layout(); fig.savefig(REP / "charts" / "s3_vram.png"); plt.close(fig)

    # Slide 4 — chất lượng + bộ ba RAG
    grouped("acc_V2", "Độ chính xác giải toán (Vấn đề 2)", "%", "s4_acc_v2.png")
    fig, ax = plt.subplots(figsize=(max(7, 2.2 * len(keys)), 3.8))
    mets = [("rag_context_relevance", "Context\nRelevance"),
            ("rag_groundedness", "Groundedness"), ("rag_answer_relevance", "Answer\nRelevance")]
    x = np.arange(len(mets))
    # Bề rộng & vị trí phải suy ra từ SỐ MODEL, không đặt cứng: ban đầu code viết
    # cho đúng 3 model (w=.26, lệch (i-1)) nên thêm model là cụm cột lệch khỏi tâm.
    w = 0.8 / len(keys)
    for i, k in enumerate(keys):
        v = [(summ.get((k, "rag"), {}) or {}).get(m) or 0 for m, _ in mets]
        b = ax.bar(x + (i - (len(keys) - 1) / 2) * w, v, w, label=labels[i])
        ax.bar_label(b, fmt="{:.0f}", fontsize=6)
    ax.set_xticks(x); ax.set_xticklabels([l for _, l in mets], fontsize=8)
    ax.set_title("Bộ ba chỉ số RAG (RAG Triad)"); ax.set_ylabel("%"); ax.legend(fontsize=7)
    fig.tight_layout(); fig.savefig(REP / "charts" / "s4_rag_triad.png"); plt.close(fig)

    # Slide 5 — prompt thuần vs RAG, cả 3 vấn đề
    for prob, nm in (("V1", "Ra đề"), ("V2", "Giải toán"), ("V3", "Chấm bài")):
        grouped(f"acc_{prob}", f"Prompt thuần vs RAG — {nm} (Vấn đề {prob[1]})", "%",
                f"s5_acc_{prob}.png")

    # Slide 7 — model được chọn, theo từng vấn đề
    ch = config.JUDGE_KEY
    fig, ax = plt.subplots(figsize=(7, 3.7))
    probs = ["V1", "V2", "V3"]; x = np.arange(3); w = .36
    for i, mode in enumerate(("pure", "rag")):
        v = [(summ.get((ch, mode), {}) or {}).get(f"acc_{p}") or 0 for p in probs]
        b = ax.bar(x + (i - .5) * w, v, w, color=C[mode],
                   label="Prompt thuần" if mode == "pure" else "Prompt + RAG")
        ax.bar_label(b, fmt="{:.1f}", fontsize=8)
    ax.set_xticks(x); ax.set_xticklabels(["Vấn đề 1\nRa đề", "Vấn đề 2\nGiải toán",
                                          "Vấn đề 3\nChấm bài"], fontsize=8)
    ax.set_title(f"{dict(zip(keys, labels))[ch]} — prompt thuần vs RAG theo từng vấn đề")
    ax.set_ylabel("%"); ax.legend(fontsize=8)
    fig.tight_layout(); fig.savefig(REP / "charts" / "s7_chosen_by_problem.png"); plt.close(fig)

    # BỘ DỄ vs BỘ KHÓ — cho thấy bộ dễ bị trần và bộ khó mới phân biệt được model.
    has_h = any((summ.get((k, m), {}) or {}).get("n_V2H") for k in keys for m in ("pure", "rag"))
    if has_h:
        fig, ax = plt.subplots(figsize=(max(8, 2.0 * len(keys)), 4.0))
        x = np.arange(len(keys)); w = 0.2
        series = [("acc_V2", "pure", "Bộ gốc · prompt thuần", "#cbd5e1"),
                  ("acc_V2", "rag", "Bộ gốc · RAG", "#64748b"),
                  ("acc_V2H", "pure", "Bộ KHÓ · prompt thuần", "#fca5a5"),
                  ("acc_V2H", "rag", "Bộ KHÓ · RAG", "#dc2626")]
        for i, (met, mode, lab, col) in enumerate(series):
            v = [(summ.get((k, mode), {}) or {}).get(met) or 0 for k in keys]
            b = ax.bar(x + (i - 1.5) * w, v, w, label=lab, color=col)
            ax.bar_label(b, fmt="%.0f", fontsize=6)
        ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=8)
        ax.set_ylabel("Độ chính xác (%)"); ax.set_ylim(0, 105)
        ax.set_title("Vấn đề 2 — bộ gốc (bị trần) vs bộ khó (phân biệt được)")
        ax.legend(fontsize=7, ncol=2)
        fig.tight_layout(); fig.savefig(REP / "charts" / "s_easy_vs_hard.png"); plt.close(fig)

    # ĐƯỜNG CONG CHẤT LƯỢNG THEO KÍCH THƯỚC — trả lời "sao không chọn 14B?".
    # Chỉ vẽ họ reasoning (cùng dòng R1-Distill) để so đúng một biến: số tham số.
    rs = [m for m in config.MODELS if m.family == "reasoning"]
    if len(rs) >= 3:
        fig, (a1, a2) = plt.subplots(1, 2, figsize=(11, 3.9))
        xs = [m.params_b for m in rs]
        for prob, mk, cl in (("V2", "o-", "#2563eb"), ("V3", "s-", "#059669"),
                             ("V1", "^-", "#d97706")):
            ys = [(summ.get((m.key, "rag"), {}) or {}).get(f"acc_{prob}") for m in rs]
            if all(y is not None for y in ys):
                a1.plot(xs, ys, mk, color=cl, label=f"{prob}", lw=2, ms=7)
                for x, y in zip(xs, ys):
                    a1.annotate(f"{y:.0f}", (x, y), textcoords="offset points",
                                xytext=(0, 7), fontsize=7, ha="center")
        a1.set_xlabel("Số tham số (tỷ)"); a1.set_ylabel("Độ chính xác (%)")
        a1.set_title("Chất lượng theo kích thước (họ R1-Distill, chế độ RAG)")
        a1.set_xticks(xs); a1.set_xticklabels([f"{m.short}\n{m.params_b}B" for m in rs], fontsize=7)
        a1.legend(fontsize=7)
        # Cái giá phải trả: độ trễ tăng theo kích thước
        lat = [(summ.get((m.key, "pure"), {}) or {}).get("latency_s_p50") or 0 for m in rs]
        a2.plot(xs, lat, "D-", color="#dc2626", lw=2, ms=7)
        for x, y in zip(xs, lat):
            a2.annotate(f"{y:.0f}s", (x, y), textcoords="offset points",
                        xytext=(0, 7), fontsize=7, ha="center")
        a2.set_xlabel("Số tham số (tỷ)"); a2.set_ylabel("Độ trễ P50 (giây)")
        a2.set_title("Cái giá: độ trễ theo kích thước (concurrency=1)")
        a2.set_xticks(xs); a2.set_xticklabels([f"{m.short}\n{m.params_b}B" for m in rs], fontsize=7)
        fig.tight_layout(); fig.savefig(REP / "charts" / "s_scaling.png"); plt.close(fig)

    # Rò chữ Trung/Nhật/Hàn — lỗi chất lượng nhìn thấy được với người dùng Việt.
    fig, ax = plt.subplots(figsize=(max(7, 1.9 * len(keys)), 3.7))
    x = np.arange(len(keys)); w = 0.36
    v1 = [(summ.get((k, "rag"), {}) or {}).get("cjk_leak_rate") or 0 for k in keys]
    v2v = [(summ.get((k, "rag"), {}) or {}).get("cjk_leak_answer_rate") or 0 for k in keys]
    b1 = ax.bar(x - w / 2, v1, w, label="Rò ở bất kỳ đâu (kể cả <think>)", color="#fbbf24")
    b2 = ax.bar(x + w / 2, v2v, w, label="Rò vào ĐÁP ÁN (sinh viên đọc thấy)", color="#dc2626")
    ax.bar_label(b1, fmt="%.1f", fontsize=7); ax.bar_label(b2, fmt="%.1f", fontsize=7)
    ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=8)
    ax.set_title("Tỉ lệ câu trả lời bị lẫn chữ Trung/Nhật/Hàn")
    ax.set_ylabel("% số câu"); ax.legend(fontsize=7)
    fig.tight_layout(); fig.savefig(REP / "charts" / "s_cjk_leak.png"); plt.close(fig)
    # Ghi chú: biểu đồ này vẽ chế độ RAG; bảng trong báo cáo dùng số GỘP hai chế độ.
    # Hai số chỉ lệch nhẹ (rò CJK gần như không phụ thuộc RAG) — xem bảng để lấy số chuẩn.

    # Phân loại lỗi
    fig, ax = plt.subplots(figsize=(max(7, 1.9 * len(keys)), 3.7))
    # Bộ KHÓ, khớp với bảng Slide 7. Bộ gốc không dùng được: 32B đạt 100% -> không có
    # câu sai nào -> cột rỗng, biểu đồ vô nghĩa.
    cats = [("errH_retrieval", "Lỗi RAG"), ("errH_calculation", "Lỗi tính toán"),
            ("errH_hallucination", "Ảo giác")]
    bot = np.zeros(len(keys))
    for (c, nm), col in zip(cats, ["#f59e0b", "#ef4444", "#8b5cf6"]):
        v = np.array([(summ.get((k, "rag"), {}) or {}).get(c) or 0 for k in keys])
        ax.bar(labels, v, bottom=bot, label=nm, color=col); bot += v
    ax.set_title("Cơ cấu nguyên nhân lỗi (Vấn đề 2 — BỘ KHÓ, chế độ RAG)")
    ax.set_ylabel("% trên tổng số câu sai"); ax.legend(fontsize=8)
    fig.tight_layout(); fig.savefig(REP / "charts" / "s_errors.png"); plt.close(fig)


def excel(summ, cost_rows, cre):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); wb.remove(wb.active)
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E78")

    def sheet(name, rows, cols=None):
        ws = wb.create_sheet(name[:31])
        if not rows:
            return
        cols = cols or list(rows[0].keys())
        ws.append(cols)
        for c in ws[1]:
            c.font = hdr; c.fill = fill
        for r in rows:
            ws.append([r.get(c) for c in cols])
        ws.freeze_panes = "A2"
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max(len(str(c)) + 2, 11), 34)

    # Bảng chi tiết theo yêu cầu: ID | Model | TTFT | Speed | VRAM | Đúng/Sai | Đánh giá
    detail = []
    for m in config.MODELS:
        for mode in ("pure", "rag"):
            g = OUT / f"gpu_{m.key}_{mode}_perf.json"
            vram = json.loads(g.read_text())["vram_peak"] / 1024 if g.exists() else None
            for r in load(m.key, mode):
                detail.append({
                    "Câu hỏi ID": r.get("task_id"), "Vấn đề": r.get("problem"),
                    "Model": m.label, "Chế độ": "Prompt thuần" if mode == "pure" else "Prompt+RAG",
                    "Môn": r.get("subject"), "Chủ đề": r.get("topic"),
                    "Độ khó": r.get("difficulty"),
                    "TTFT (ms)": round(r["ttft_ms"], 1) if r.get("ttft_ms") else None,
                    "Speed (tok/s)": round(r["tokens_per_sec"], 1) if r.get("tokens_per_sec") else None,
                    "Latency (s)": round(r.get("total_latency_s", 0), 2),
                    "VRAM (GB)": round(vram, 1) if vram else None,
                    "Think (s)": round(r["think_time_s"], 2) if r.get("think_time_s") else None,
                    "Token nghĩ": r.get("think_tokens"), "Token đáp": r.get("answer_tokens"),
                    "Kết quả": ("Đúng" if r.get("correct") else "Sai") if "correct" in r else "—",
                    "Cách chấm": r.get("graded"),
                    "Nguyên nhân lỗi": r.get("error_category"),
                    "Context đủ?": r.get("context_sufficient"),
                    "Bám tài liệu?": r.get("grounded"),
                    "Đúng định dạng?": r.get("format_ok"),
                    "Bị cắt?": r.get("truncated"),
                    "Đáp án model": (r.get("final_answer") or "")[:200],
                    "Lỗi": (r.get("error") or "")[:120],
                })
    sheet("Chi tiết từng câu", detail)
    sheet("Tổng hợp", [{"model": k[0], "mode": k[1], **v} for k, v in summ.items()])
    sheet("Chi phí", cost_rows)
    cal = OUT / "judge_calibration.json"
    if cal.exists():
        c = json.loads(cal.read_text(encoding="utf-8"))
        sheet("Hiệu chuẩn giám khảo", c["detail"])
    sheet("Cấu hình model", [{
        "Model": m.label, "HF ID": m.hf_id, "Loại": m.family, "Tham số (tỷ)": m.params_b,
        "Temperature": m.temperature, "Top-P": m.top_p, "Max tokens": m.max_tokens,
        "Lý do chọn tham số": m.temp_rationale} for m in config.MODELS])
    wb.save(REP / "ket_qua_benchmark.xlsx")


def main():
    print("Tính Context Relevance bằng embedding...")
    try:
        cre = context_relevance_embedding()
    except Exception as e:  # noqa: BLE001
        print("  bỏ qua (bge-m3 không sẵn sàng):", e); cre = {}
    summ = {}
    for m in config.MODELS:
        for mode in ("pure", "rag"):
            rows = load(m.key, mode)
            if rows:
                # _model/_mode để report.py tra ngược được file gpu_*.json tương ứng.
                summ[(m.key, mode)] = {"_model": m.key, "_mode": mode,
                                       **summarize(rows, cre), **cjk_leak(m.key, mode)}
    if not summ:
        print("Chưa có dữ liệu đã chấm."); return
    cost_rows = costs(summ)
    (OUT / "summary.json").write_text(
        json.dumps({f"{k[0]}|{k[1]}": v for k, v in summ.items()}, ensure_ascii=False, indent=1),
        encoding="utf-8")
    charts(summ)
    excel(summ, cost_rows, cre)
    print("Đã ghi:", REP / "ket_qua_benchmark.xlsx")
    for (m, mode), d in summ.items():
        print(f"  {m:8s} {mode:5s} | V2 {d['acc_V2']:5.1f}% | V1 {d['acc_V1']:5.1f}% | "
              f"V3 {d['acc_V3']:5.1f}% | TTFT p50 {d.get('ttft_ms_p50')} ms | "
              f"{d.get('tokens_per_sec_mean')} tok/s")


if __name__ == "__main__":
    main()
